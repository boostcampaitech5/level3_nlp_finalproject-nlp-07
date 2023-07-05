from bs4 import BeautifulSoup as bs
from pathlib import Path
from typing import Optional,Union,Dict,List
from openpyxl import Workbook
import time
import os
import re
import requests as rq
import json
import csv
from datetime import datetime
from pytz import timezone
def get_headers(
    key: str,
    default_value: Optional[str] = None
    )-> Dict[str,Dict[str,str]]:
    """ Get Headers """
    JSON_FILE : str = 'json/headers.json'

    with open(JSON_FILE,'r',encoding='UTF-8') as file:
        headers : Dict[str,Dict[str,str]] = json.loads(file.read())

    try :
        return headers[key]
    except:
        if default_value:
            return default_value
        raise EnvironmentError(f'Set the {key}')

class Coupang:
    @staticmethod
    def get_product_code(url: str)-> str:
        """ 입력받은 URL 주소의 PRODUCT CODE 추출하는 메소드 """
        prod_code : str = url.split('products/')[-1].split('?')[0]
        return prod_code

    def __init__(self)-> None:
        self.__headers : Dict[str,str] = get_headers(key='headers')

    def main(self, url_list: List[str]) -> List[List[Dict[str, Union[str, int]]]]:
        # 각 URL의 첫 페이지에서 리뷰를 가져옴
        result = []
        for URL in url_list:
            # URL의 Product Code 추출
            prod_code: str = self.get_product_code(url=URL)

            for page in range(1, 30):
                # URL 주소 재가공
                url_to_fetch = f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={page}&size=10&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true'

                # __headers에 referer 키 추가
                self.__headers['referer'] = URL

                with rq.Session() as session:
                    page_data = self.fetch(url=url_to_fetch, session=session)
                    result.append(page_data)

        return result

    def fetch(self,url:str,session)-> List[Dict[str,Union[str,int]]]:
        save_data : List[Dict[str,Union[str,int]]] = list()

        with session.get(url=url,headers=self.__headers) as response :
            html = response.text
            soup = bs(html,'html.parser')

            # Article Boxes
            article_lenth = len(soup.select('article.sdp-review__article__list'))

            for idx in range(article_lenth):
                dict_data : Dict[str,Union[str,int]] = dict()
                articles = soup.select('article.sdp-review__article__list')

                # 구매자 이름
                user_name = articles[idx].select_one('span.sdp-review__article__list__info__user__name')
                if user_name == None or user_name.text == '':
                    user_name = '-'
                else:
                    user_name = user_name.text.strip()

                # 평점
                rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange')
                if rating == None:
                    rating = 0
                else :
                    rating = int(rating.attrs['data-rating'])

                # 구매자 상품명
                prod_name = articles[idx].select_one('div.sdp-review__article__list__info__product-info__name')
                if prod_name == None or prod_name.text == '':
                    prod_name = '-'
                else:
                    prod_name = prod_name.text.strip()

                # 헤드라인(타이틀)
                headline = articles[idx].select_one('div.sdp-review__article__list__headline')
                if headline == None or headline.text == '':
                    headline = ''
                else:
                    headline = headline.text.strip()

                # 리뷰 내용
                review_content = articles[idx].select_one('div.sdp-review__article__list__review > div')
                if review_content == None :
                    review_content = ''
                else:
                    review_content = re.sub('[\n\t]','',review_content.text.strip())

                # 맛 만족도
                answer = articles[idx].select_one('span.sdp-review__article__list__survey__row__answer')
                if answer == None or answer.text == '':
                    answer = ''
                else:
                    answer = answer.text.strip()


                # 원본 URL
                dict_data['url'] = url
                dict_data['prod_name'] = prod_name
                dict_data['user_name'] = user_name
                dict_data['rating'] = rating
                dict_data['headline'] = headline
                dict_data['review_content'] = review_content
                dict_data['answer'] = answer

                save_data.append(dict_data)

                print(dict_data , '\n')

            return save_data

    def input_review_url(self)-> str:
        while True:
            # Window
            # os.system('cls')
            # Mac
            os.system('clear')
            
            # Review URL
            review_url : str = input('원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906&q=%ED%9E%98%EB%82%B4%EB%B0%94+%EC%B4%88%EC%BD%94+%EC%8A%A4%EB%8B%88%EC%BB%A4%EC%A6%88&itemsCount=36&searchId=0c5c84d537bc41d1885266961d853179&rank=2&isAddedCart=\n\n:')
            if not review_url :
                # Window
                os.system('cls')
                # Mac
                #os.system('clear')
                print('URL 주소가 입력되지 않았습니다')
                continue
            return review_url

    def input_page_count(self)-> int:
        # Window
        # os.system('cls')
        # Mac
        os.system('clear')
        while True:
            page_count : str = input('페이지 수를 입력하세요\n\n:')
            if not page_count:
                print('페이지 수가 입력되지 않았습니다\n')
                continue

            return int(page_count)
        
def load_urls(file_path: str) -> List[str]:
    # Load URLs from a CSV file
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        urls = [row['URL'] for row in reader]
    return urls





class OpenPyXL:
    @staticmethod
    def save_file()-> None:
        # 크롤링 결과
        results : List[List[Dict[str,Union[str,int]]]] = Coupang().main()

        wb = Workbook()
        ws = wb.active
        ws.append(['상품명','구매자 이름','구매자 평점','리뷰 제목','리뷰 내용','맛 만족도'])

        row = 2

        for x in results:
            for result in x :
                ws[f'A{row}'] = result['prod_name']
                ws[f'B{row}'] = result['user_name']
                ws[f'C{row}'] = result['rating']
                ws[f'D{row}'] = result['headline']
                ws[f'E{row}'] = result['review_content']
                ws[f'F{row}'] = result['answer']

                row += 1

        savePath : str = os.path.abspath('쿠팡-상품리뷰-크롤링')
        fileName : str = results[0][0]['prod_name'] + '.xlsx'

        if not os.path.exists(savePath):
            os.mkdir(savePath)

        wb.save(os.path.join(savePath,fileName))
        wb.close()

        print(f'파일 저장완료!\n\n{os.path.join(savePath,fileName)}')


class CSV:
    @staticmethod
    def save_file()-> None:

        product_name = '떡볶이'

        file_path = f"./data/products/{product_name}.csv"
        url_list = load_urls(file_path)

        # 크롤링 결과
        results : List[List[Dict[str,Union[str,int]]]] = Coupang().main(url_list)

        # 파일에 쓸 데이터 준비
        csv_columns = ['prod_name', 'user_name', 'rating', 'headline', 'review_content', 'answer', 'url']
        # 서울 시간
        now = datetime.now(timezone('Asia/Seoul'))
        # 파일 이름
        # csv_file = f'../reviews/{product_name}_{now.strftime("%y%m%d_%H")}.csv'
        csv_file = f'./data/reviews/{product_name}리뷰.csv'

        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=csv_columns)
            writer.writeheader()
            for data in results:
                for item in data:
                    writer.writerow(item)

        print(f'파일 저장완료!\n\n{csv_file}')


if __name__ == '__main__':

    # OpenPyXL.save_file()
    CSV.save_file()
